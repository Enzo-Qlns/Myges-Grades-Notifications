import openpyxl
from openpyxl.styles import Font


class ExcelService:
    def __init__(self, filename):
        self.filename = filename
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Grades"

    def save_grades(self, grades):
        """
        Enregistre les notes dans le fichier Excel.
        :param grades: Liste de dictionnaires contenant les notes.
        :return: None
        """
        # Ajouter les en-têtes
        headers = [
            "Course", "Code", "Grades", "Bonus", "Exam", "Average", "Trimester",
            "Trimester Name", "Year", "RC ID", "ECTS", "Coef", "Teacher Civility",
            "Teacher First Name", "Teacher Last Name", "Absences", "Lates",
            "Letter Mark", "CC Average", "Links"
        ]
        self.sheet.append(headers)
        for cell in self.sheet[1]:
            cell.font = Font(bold=True)

        # Ajouter les données des notes
        for grade in grades:
            self.sheet.append([
                grade["course"], grade["code"], ", ".join(map(str, grade["grades"])),
                grade["bonus"], grade["exam"], grade["average"], grade["trimester"],
                grade["trimester_name"], grade["year"], grade["rc_id"], grade["ects"],
                grade["coef"], grade["teacher_civility"], grade["teacher_first_name"],
                grade["teacher_last_name"], grade["absences"], grade["lates"],
                grade["letter_mark"], grade["ccaverage"], ", ".join(grade["links"])
            ])

        # Sauvegarder le fichier
        self.workbook.save(self.filename)

    def get_grades(self):
        """
        Récupère les notes du fichier Excel.
        :return: Liste de dictionnaires contenant les notes.
        """
        grades = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            grades.append({
                "course": row[0],
                "code": row[1],
                "grades": row[2].split(", ") if row[2] else [],
                "bonus": row[3],
                "exam": row[4],
                "average": row[5],
                "trimester": row[6],
                "trimester_name": row[7],
                "year": row[8],
                "rc_id": row[9],
                "ects": row[10],
                "coef": row[11],
                "teacher_civility": row[12],
                "teacher_first_name": row[13],
                "teacher_last_name": row[14],
                "absences": row[15],
                "lates": row[16],
                "letter_mark": row[17],
                "ccaverage": row[18],
                "links": row[19].split(", ") if row[19] else []
            })
        return grades

    def delete_all_grades(self):
        """
        Supprime toutes les notes du fichier Excel.
        :return: None
        """
        self.sheet.delete_rows(2, self.sheet.max_row)
        self.workbook.save(self.filename)
